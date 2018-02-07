#! python3
#! -*- coding: utf-8 -*-
import openpyxl

# 合并文件列表
fileList = [
  '/Users/it/Desktop/2.xlsx',
  '/Users/it/Desktop/3.xlsx'
]
# 跳过表头
skipHead = True

if __name__ == '__main__':
  newWorkBook = openpyxl.Workbook()
  newWorkBook.create_sheet()
  newWorkBook.save('merged.xlsx')
  for file in fileList:
    print('读取文件:' + file)
    wb = openpyxl.load_workbook(file).active
    print('读取 sheet:' + wb.title)
    for row in wb.values:
      if skipHead and tuple(row) == tuple(wb.values)[0]:
        print('跳过表头:' + str(tuple(row)))
        continue
      print(row)
      newWorkBook.active.append(row)
    newWorkBook.save('merged.xlsx')
    

